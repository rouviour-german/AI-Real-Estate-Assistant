import tempfile
from pathlib import Path

import pandas as pd
import pytest

from data.csv_loader import DataLoaderCsv, DataLoaderExcel
from yarl import URL


def test_convert_github_url_to_raw():
    u = "https://github.com/user/repo/blob/main/data.csv"
    r = DataLoaderCsv.convert_github_url_to_raw(u)
    assert r == "https://raw.githubusercontent.com/user/repo/main/data.csv"


def test_format_df_adds_missing_columns_and_normalizes():
    df = pd.DataFrame(
        {
            "City": ["Krakow", None],
            "Rooms": [2, None],
            "Price": [900, 1200],
            "has_parking": ["yes", None],
            "has_balcony": [1, 0],
        }
    )
    out = DataLoaderCsv.format_df(df)
    assert "city" in out.columns
    assert "rooms" in out.columns
    assert "bathrooms" in out.columns
    assert out.loc[0, "has_parking"] in (True, False)
    assert isinstance(out.loc[0, "rooms"], float)
    assert len(out) == 2


def test_bathrooms_fake_logic():
    assert DataLoaderCsv.bathrooms_fake(1.0) == 1.0
    v = DataLoaderCsv.bathrooms_fake(3.0)
    assert v in (1.0, 2.0)


def test_load_df_reads_csv_from_path():
    df_in = pd.DataFrame({"city": ["Krakow"], "price": [900], "rooms": [2]})
    with tempfile.TemporaryDirectory() as tmp:
        p = Path(tmp) / "data.csv"
        df_in.to_csv(p, index=False)
        loader = DataLoaderCsv(p)
        df_out = loader.load_df()

    assert list(df_out.columns) == ["city", "price", "rooms"]
    assert len(df_out) == 1


def test_load_df_reads_excel_from_path():
    pytest.importorskip("openpyxl")

    df_in = pd.DataFrame({"city": ["Warsaw"], "price": [1200], "rooms": [3]})
    with tempfile.TemporaryDirectory() as tmp:
        p = Path(tmp) / "data.xlsx"
        df_in.to_excel(p, index=False)
        loader = DataLoaderCsv(p)
        df_out = loader.load_df()

    assert list(df_out.columns) == ["city", "price", "rooms"]
    assert len(df_out) == 1


def test_excel_get_sheet_names_xlsx(monkeypatch):
    class Workbook:
        sheetnames = ["Sheet1", "Sheet2"]

        def close(self):
            return None

    import openpyxl

    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_args, **_kwargs: Workbook())
    loader = DataLoaderExcel("fake.xlsx")
    assert loader.get_sheet_names() == ["Sheet1", "Sheet2"]


def test_excel_get_sheet_names_xls(monkeypatch):
    class Workbook:
        def sheet_names(self):
            return ["SheetA"]

    import xlrd

    monkeypatch.setattr(xlrd, "open_workbook", lambda *_args, **_kwargs: Workbook())
    loader = DataLoaderExcel("fake.xls")
    assert loader.get_sheet_names() == ["SheetA"]


def test_excel_get_sheet_names_ods_fallback(monkeypatch):
    class ExcelFile:
        sheet_names = ["ODS1"]

        def close(self):
            return None

    import odf.opendocument

    monkeypatch.setattr(
        odf.opendocument, "load", lambda *_args, **_kwargs: (_ for _ in ()).throw(ImportError())
    )
    monkeypatch.setattr(pd, "ExcelFile", lambda *_args, **_kwargs: ExcelFile())
    loader = DataLoaderExcel("fake.ods")
    assert loader.get_sheet_names() == ["ODS1"]


def test_excel_get_sheet_names_other_suffix(monkeypatch):
    class ExcelFile:
        sheet_names = ["MacroSheet"]

        def close(self):
            return None

    monkeypatch.setattr(pd, "ExcelFile", lambda *_args, **_kwargs: ExcelFile())
    loader = DataLoaderExcel("fake.xlsm")
    assert loader.get_sheet_names() == ["MacroSheet"]


def test_excel_load_df_builds_kwargs(monkeypatch):
    captured = {}

    def fake_read_excel(_path, **kwargs):
        captured.update(kwargs)
        return pd.DataFrame({"city": ["Wroclaw"], "price": [1000], "rooms": [2]})

    monkeypatch.setattr(pd, "read_excel", fake_read_excel)
    loader = DataLoaderExcel("fake.xlsx", sheet_name="Sheet1", header_row=1)
    df_out = loader.load_df()
    assert captured["engine"] == "openpyxl"
    assert captured["sheet_name"] == "Sheet1"
    assert captured["header"] == 1
    assert len(df_out) == 1


def test_detect_source_type_with_url_instance():
    url = URL("https://example.com/data.xlsx")
    assert DataLoaderExcel.detect_source_type(url) == "url"
